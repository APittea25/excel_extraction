import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from io import BytesIO
import re
import os
from collections import defaultdict
import graphviz

st.set_page_config(page_title="Named Range Formula Remapper", layout="wide")
st.title("\U0001F4D8 Named Range Coordinates + Formula Remapping")

# --- Toggle button ---
if "expanded_all" not in st.session_state:
    st.session_state.expanded_all = False

if st.button("🔁 Expand/Collapse All Named Ranges"):
    st.session_state.expanded_all = not st.session_state.expanded_all

# --- Inject JS for expand/collapse ---
if st.session_state.expanded_all:
    st.markdown("""
    <script>
    document.querySelectorAll('details').forEach(el => { el.open = true; });
    </script>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <script>
    document.querySelectorAll('details').forEach(el => { el.open = false; });
    </script>
    """, unsafe_allow_html=True)

# Allow manual mapping of external references like [1], [2], etc.
st.subheader("Manual Mapping for External References")
external_refs = {}
for i in range(1, 10):
    ref_key = f"[{i}]"
    workbook_name = st.text_input(f"Map external reference {ref_key} to workbook name (e.g., Mortality_Model_Inputs.xlsx)", key=ref_key)
    if workbook_name:
        external_refs[ref_key] = workbook_name

uploaded_files = st.file_uploader("\U0001F4C2 Upload Excel files", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    all_named_cell_map = {}
    all_named_ref_info = {}
    file_display_names = {}
    named_ref_formulas = {}

    for uploaded_file in uploaded_files:
        display_name = uploaded_file.name
        file_display_names[display_name] = uploaded_file
        st.header(f"\U0001F4C4 File: `{display_name}`")
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=False)

        for name in wb.defined_names:
            dn = wb.defined_names[name]
            if dn.is_external or not dn.attr_text:
                continue
            for sheet_name, ref in dn.destinations:
                try:
                    ws = wb[sheet_name]
                    ref_clean = ref.replace("$", "").split("!")[-1]
                    cells = ws[ref_clean] if ":" in ref_clean else [[ws[ref_clean]]]

                    min_row = min(cell.row for row in cells for cell in row)
                    min_col = min(cell.column for row in cells for cell in row)

                    coord_set = set()
                    for row in cells:
                        for cell in row:
                            r, c = cell.row, cell.column
                            row_offset = r - min_row + 1
                            col_offset = c - min_col + 1
                            all_named_cell_map[(display_name, sheet_name, r, c)] = (name, row_offset, col_offset)
                            coord_set.add((r, c))
                    all_named_ref_info[name] = (display_name, sheet_name, coord_set, min_row, min_col)
                except:
                    continue

    def remap_formula(formula, current_file, current_sheet):
        if not formula:
            return ""

        def cell_address(row, col):
            return f"{get_column_letter(col)}{row}"

        def remap_single_cell(ref, default_file, default_sheet):
            if "!" in ref:
                sheet_part, addr = ref.split("!")
                match = re.match(r"\[(\d+)\]", sheet_part)
                if match:
                    external_ref = match.group(0)
                    external_file = external_refs.get(external_ref, external_ref)
                    return f"[{external_file}]{ref}"
                sheet_name = sheet_part
            else:
                sheet_name = default_sheet
                addr = ref

            addr = addr.replace("$", "").upper()
            match = re.match(r"([A-Z]+)([0-9]+)", addr)
            if not match:
                return ref
            col_str, row_str = match.groups()
            row = int(row_str)
            col = column_index_from_string(col_str)

            key = (default_file, sheet_name, row, col)
            if key in all_named_cell_map:
                name, r_off, c_off = all_named_cell_map[key]
                return f"[{default_file}]{name}[{r_off}][{c_off}]"
            else:
                return f"[{default_file}]{sheet_name}!{addr}"

        def remap_range(ref, default_file, default_sheet):
            if ref.startswith("["):
                match = re.match(r"\[(\d+)\]", ref)
                if match:
                    external_ref = match.group(0)
                    external_file = external_refs.get(external_ref, external_ref)
                    return f"[{external_file}]{ref}"

            if "!" in ref:
                sheet_name, addr = ref.split("!")
            else:
                sheet_name = default_sheet
                addr = ref

            addr = addr.replace("$", "").upper()
            if ":" not in addr:
                return remap_single_cell(ref, default_file, default_sheet)

            start, end = addr.split(":")
            m1 = re.match(r"([A-Z]+)([0-9]+)", start)
            m2 = re.match(r"([A-Z]+)([0-9]+)", end)
            if not m1 or not m2:
                return ref
            start_col = column_index_from_string(m1.group(1))
            start_row = int(m1.group(2))
            end_col = column_index_from_string(m2.group(1))
            end_row = int(m2.group(2))

            label_set = set()
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    key = (default_file, sheet_name, row, col)
                    if key in all_named_cell_map:
                        name, r_off, c_off = all_named_cell_map[key]
                        label_set.add(f"[{default_file}]{name}[{r_off}][{c_off}]")
                    else:
                        label_set.add(f"[{default_file}]{sheet_name}!{cell_address(row, col)}")
            return ", ".join(sorted(label_set))

        pattern = r"(?<![A-Za-z0-9_])(?:\[[^\]]+\])?[A-Za-z0-9_]+!\$?[A-Z]{1,3}\$?[0-9]{1,7}(?::\$?[A-Z]{1,3}\$?[0-9]{1,7})?|(?<![A-Za-z0-9_])\$?[A-Z]{1,3}\$?[0-9]{1,7}(?::\$?[A-Z]{1,3}\$?[0-9]{1,7})?"
        matches = list(re.finditer(pattern, formula))
        replaced_formula = formula
        offset = 0
        for match in matches:
            raw = match.group(0)
            remapped = remap_range(raw, current_file, current_sheet)
            start, end = match.start() + offset, match.end() + offset
            replaced_formula = replaced_formula[:start] + remapped + replaced_formula[end:]
            offset += len(remapped) - len(raw)
        return replaced_formula

    for (name, (file_name, sheet_name, coord_set, min_row, min_col)) in all_named_ref_info.items():
        entries = []
        formulas_for_graph = []

        try:
            file_bytes = file_display_names[file_name]
            wb = load_workbook(BytesIO(file_bytes.getvalue()), data_only=False)
            ws = wb[sheet_name]
            min_col_letter = get_column_letter(min([c for (_, c) in coord_set]))
            max_col_letter = get_column_letter(max([c for (_, c) in coord_set]))
            min_row_num = min([r for (r, _) in coord_set])
            max_row_num = max([r for (r, _) in coord_set])
            ref_range = f"{min_col_letter}{min_row_num}:{max_col_letter}{max_row_num}"
            cell_range = ws[ref_range] if ":" in ref_range else [[ws[ref_range]]]

            for row in cell_range:
                for cell in row:
                    row_offset = cell.row - min_row + 1
                    col_offset = cell.column - min_col + 1
                    label = f"{name}[{row_offset}][{col_offset}]"

                    try:
                        formula = None
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula = cell.value.strip()
                        elif hasattr(cell, 'value') and hasattr(cell.value, 'text'):
                            formula = str(cell.value.text).strip()
                        elif hasattr(cell, 'value'):
                            formula = str(cell.value)

                        if formula:
                            remapped = remap_formula(formula, file_name, sheet_name)
                            formulas_for_graph.append(remapped)
                        elif cell.value is not None:
                            formula = f"[value] {str(cell.value)}"
                            remapped = formula
                        else:
                            formula = "(empty)"
                            remapped = formula
                    except Exception as e:
                        formula = f"[error reading cell: {e}]"
                        remapped = formula

                    entries.append(f"{label} = {formula}\n → {remapped}")
        except Exception as e:
            entries.append(f"❌ Error accessing `{name}` in `{sheet_name}`: {e}")

        named_ref_formulas[name] = formulas_for_graph

        with st.expander(f"📌 Named Range: `{name}` → `{sheet_name}` in `{file_name}`"):
            st.code("\n".join(entries), language="text")

    # Dependency Graph
    st.subheader("🔗 Dependency Graph")
    dot = graphviz.Digraph()
    dot.attr(compound='true', rankdir='LR')

    grouped = defaultdict(list)
    for name, (file, *_rest) in all_named_ref_info.items():
        grouped[file].append(name)

    dependencies = defaultdict(set)
    for target, formulas in named_ref_formulas.items():
        joined = " ".join(formulas)
        for source in named_ref_formulas:
            if source != target and re.search(rf"\b{re.escape(source)}\b", joined):
                dependencies[target].add(source)

    for i, (file_name, nodes) in enumerate(grouped.items()):
        with dot.subgraph(name=f"cluster_{i}") as c:
            c.attr(label=file_name)
            c.attr(style='filled', color='lightgrey')
            for node in nodes:
                c.node(node)

    for target, sources in dependencies.items():
        for source in sources:
            dot.edge(source, target)

    st.graphviz_chart(dot)
else:
    st.info("⬆️ Upload one or more `.xlsx` files to begin.")
